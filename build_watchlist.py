"""
Parse the analyst coverage workbook into watchlist.csv.

Hierarchy is encoded by the header cell's fill THEME color:
    theme == 3  -> top-level GROUP   (e.g. 'China Internet')
    theme == 0  -> SUB-GROUP         (e.g. 'ERP' under 'China Software')
A row with a ticker in column B is a stock; a row without is a header.

Bloomberg-style tickers ("700 HK", "BABA US", "600588 CH") are converted to
Yahoo Finance symbols. Output columns: group, subgroup, name, ticker, bbg, currency
"""
import csv
import openpyxl

SRC = r"C:\Users\jeffl\OneDrive\文档\Aqua Lake Capital - Interview\JL coverage_v3.xlsx"
OUT = r"C:\Users\jeffl\Claude project\5) Coverage-and-alert-agent\watchlist.csv"
REPORT = r"C:\Users\jeffl\Claude project\5) Coverage-and-alert-agent\watchlist_report.txt"

# Indices / non-standard symbols that don't follow the code+exchange rule.
OVERRIDES = {
    "HSTECH HK": "HSTECH.HK",   # Hang Seng TECH Index (Yahoo serves only a current level, no history)
    "931087 CH": "931087.SS",   # CSI index (Yahoo serves only a current level, no history)
    "GOTO US": "GOTO.JK",       # GoTo Gojek is Jakarta-listed (IDX), not US
}

# Currency overrides for symbols whose trading currency != the Bloomberg exchange default.
CCY_OVERRIDE = {"GOTO US": "IDR"}

CCY_BY_EXCH = {"US": "USD", "HK": "HKD", "CH": "CNY", "JP": "JPY", "TB": "THB"}

# Bloomberg tickers to drop entirely (e.g. indices with no usable Yahoo price history).
EXCLUDE = {"HSTECH HK", "931087 CH"}


def to_yahoo(bbg: str) -> str:
    bbg = bbg.strip()
    if bbg in OVERRIDES:
        return OVERRIDES[bbg]
    code, exch = bbg.split()
    if exch == "US":
        return code.upper()
    if exch == "HK":
        return f"{int(code):04d}.HK"          # zero-pad: '1 HK' -> '0001.HK'
    if exch == "JP":
        return f"{code}.T"
    if exch == "TB":
        return f"{code}.BK"
    if exch == "CH":
        # 6xxxxx (incl. 688 STAR) and 9xxxxx -> Shanghai .SS; 0/3xxxxx -> Shenzhen .SZ
        return f"{code}.SS" if code[0] in "69" else f"{code}.SZ"
    return code


def currency_of(bbg: str) -> str:
    bbg = bbg.strip()
    if bbg in CCY_OVERRIDE:
        return CCY_OVERRIDE[bbg]
    return CCY_BY_EXCH.get(bbg.split()[-1], "")


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Tracker"]

    rows = []
    group = subgroup = ""
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1)
        name = a.value
        bbg = ws.cell(row=r, column=2).value
        if name is None:
            continue
        name = str(name).strip()
        if name in ("Coverage list", "Name"):
            continue

        is_header = bbg in (None, "")
        if is_header:
            theme = getattr(a.fill.fgColor, "theme", None)
            if theme == 3:                 # top-level group
                group, subgroup = name, ""
            else:                          # sub-group (theme 0)
                subgroup = name
            continue

        bbg = str(bbg).strip()
        if bbg in EXCLUDE:
            continue
        rows.append({
            "group": group,
            "subgroup": subgroup,
            "name": name,
            "ticker": to_yahoo(bbg),
            "bbg": bbg,
            "currency": currency_of(bbg),
        })

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["group", "subgroup", "name", "ticker", "bbg", "currency"])
        w.writeheader()
        w.writerows(rows)

    # human-readable report
    lines = [f"Parsed {len(rows)} stocks.\n"]
    cur_g = cur_s = None
    for x in rows:
        if x["group"] != cur_g:
            cur_g, cur_s = x["group"], None
            lines.append(f"\n## {x['group']}")
        if x["subgroup"] != cur_s:
            cur_s = x["subgroup"]
            if cur_s:
                lines.append(f"  ~ {cur_s}")
        lines.append(f"     {x['name']:<26} {x['bbg']:<12} -> {x['ticker']:<14} [{x['currency']}]")
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Wrote {len(rows)} rows to watchlist.csv and report to watchlist_report.txt")
    groups = {}
    for x in rows:
        groups.setdefault(x["group"], 0)
        groups[x["group"]] += 1
    print("Groups:", {g: n for g, n in groups.items()})


if __name__ == "__main__":
    main()
